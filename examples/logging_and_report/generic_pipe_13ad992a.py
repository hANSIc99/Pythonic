import time, queue, datetime
from pathlib import Path
from collections import deque
from openpyxl import workbook 
from openpyxl import load_workbook
try:
    from element_types import Record, Function, ProcCMD, GuiCMD, PythonicError
except ImportError:    
    from Pythonic.element_types import Record, Function, ProcCMD, GuiCMD, PythonicError
    
class Element(Function):

    def __init__(self, id, config, inputData, return_queue, cmd_queue):
        super().__init__(id, config, inputData, return_queue, cmd_queue)


    def execute(self):


        #####################################
        #                                   #
        #         REPORT GENERATOR          #
        #                                   #
        #####################################

        path = Path.home() / 'Pythonic' / 'executables' / 'report_template.xlsx'

        try: 
            wb = load_workbook(path)
        except FileNotFoundError as e:
            recordDone = Record(PythonicError(e), 'Template not found') 
            self.return_queue.put(recordDone)
            con.close()    
            return    
        except Exception as e:
            recordDone = Record(PythonicError(e), 'Open log for details') 
            self.return_queue.put(recordDone)
            con.close()    
            return

        # Workbook https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html
        # Cell https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html

        sheets = wb.sheetnames
        datasheet = wb['Data']


        # create an iterator over the rows in the datasheet
        rows = datasheet.iter_rows(min_row=2, max_row=999, min_col=0, max_col=2)

        # Convert unix time [s] back into a datetime object, returns an iterator
        reportdata_dt = map(lambda foo: (datetime.datetime.fromtimestamp(foo[0]), foo[1]), self.inputData)



        # iterate till the first iterator is exhausted
        for (dt, val), (row_dt, row_val) in zip(reportdata_dt, rows):
            row_dt.value = dt
            row_val.value = val
        
        # alternative approach (functional)

        # zip both iterators together
        # data_it = zip(reportdata_dt, rows)
        # def write_row(data):
        #     (dt, val), (row_dt, row_val) = data
        #     row_dt.value = dt
        #     row_val.value = val

        # maker = map(write_row, data_it)

        # deque(maker, maxlen=0)

        reportDate = datetime.datetime.now().strftime('%d_%b_%Y_%H_%M_%S')
        filename = 'report_{}.xlsx'.format(reportDate)
        filepath = Path.home() / 'Pythonic' / 'log' / filename
        wb.save(filepath)
        wb.close()

        recordDone = Record(filepath, 'Report saved under: {}'.format(filename))     
        self.return_queue.put(recordDone)